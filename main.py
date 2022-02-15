from distutils import command
from html.parser import HTMLParser
from xml.etree.ElementTree import tostring

import os, sys
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import date
from datetime import datetime
import datetime
import time

import tkinter
from tkinter import *
import tkinter.ttk

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

import platform
import pyshorteners


nameCount = 1
priceCount = 1
index = 1

nowDate = date.today().isoformat()
now=datetime.datetime.now()


printTitles = ''
printPrice = ''



url = 'https://search.shopping.naver.com/best/category/click?categoryCategoryId=50000151&categoryChildCategoryId=&categoryDemo=A00&categoryMidCategoryId=50000151&categoryRootCategoryId=50000003&chartRank=1&period=P7D'

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument("headless")

def make_clickable(val):
    return '<a href="{}">{}</a>'.format(val,val)

def activeLabel(text):
    pProgress['text'] = text
        
def scanPlatform():
    scanOS = platform.system()
    if scanOS == 'Windows':
        pass
    elif scanOS == 'Darwin':
        pass


def filtering_string(value):
    finalValue = value.replace(',','').replace('무료','').replace('최저 ','').replace('원','').replace('(','').replace(')','')
    return finalValue

def doScrollDown(whileSeconds, driver):
    
    start = datetime.datetime.now()
    end = start + datetime.timedelta(seconds=whileSeconds)
    
    while True:
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        #time.sleep(1)
        if datetime.datetime.now() > end:
            break
    

def initPrototype():
    
    activeLabel('에러발생!! 나중에 다시 시도해주세요 :)')
    time.sleep(1)    
    nameCount = 1
    priceCount = 1
    deviceNameList = []
    lowPriceList = []
    reviewList = []
    driver = webdriver.Chrome(executable_path='./chromedriver', options=options)

    # driver.implicitly_wait(3)
    # url에 접근한다.
    driver.get(url)

    #주간 버튼 클릭 (일간일때는 요놈만 주석처리하거나 나중에 선택지를 주는 방법으로 선택가능)
    driver.find_element_by_xpath("//*[@id='__next']/div/div[2]/div[2]/div/div[2]/div[2]/a[2]").click()


    doScrollDown(2,driver) #2초간 스크롤 넉넉하게 2초 동안 스크롤함

    req = driver.page_source #스크롤 한 후 페이지 소스 불러오기
    soup = BeautifulSoup(req, 'html.parser')

    titles = soup.select('div.imageProduct_title__3TsP1') #제품 이름 출력
    shopLink = soup.select('a.imageProduct_link_item__2i1IN') #쇼핑 링크
    lowPrices = soup.select('div.imageProduct_price__3vXjm') #최저가
    


    print('주간 Best100 크롤링 초기 데이터 :)')

    for title in titles:
    
        printTitles = title.get_text()
    
        filterTitle = filtering_string(printTitles)
        print(nameCount)
        print(printTitles)
    
        nameCount = nameCount+1
    
        deviceNameList.append(filterTitle)
    
    
    print() 
    print("순위별 최저가")   
    

    for lowPrice in lowPrices:
    
        printPrice = lowPrice.get_text()
        filterPrice = filtering_string(printPrice)
        print(priceCount)
        print(printPrice)

        priceCount = priceCount+1
    
        lowPriceList.append(int(filterPrice))
        

    reviewIndex = 1
    for links in shopLink:
        
        
        driver.get(links["href"])
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')
        review = soup.select_one('#section_review > div.filter_sort_group__Y8HA1 > div.filter_evaluation_tap__-45pp > ul > li.filter_on__X0_Fb > a > em')
        
        if review == None:
            
            print(str(reviewIndex) +' '+'실패(-1)')
            reviewList.append(-1)
            reviewIndex = reviewIndex + 1
        else:
            reviewText = review.get_text()
            filterReviewCount = filtering_string(reviewText)
            print(str(reviewIndex) +' '+filterReviewCount)
            
            reviewList.append(int(filterReviewCount))
            reviewIndex = reviewIndex + 1
            

    
    list_of_tuples = list(zip(deviceNameList,lowPriceList,reviewList))
    
    df = pd.DataFrame(list_of_tuples, columns = ['제품명', '최저가', '리뷰수'])
    
    df.index = df.index+1
    
    df.to_excel('./'+nowDate+'.xlsx', header=True, index= True, sheet_name='초기데이터',index_label='순위')

    driver.quit()
    
    activeLabel('완료!')



def bestRank():
    
    activeLabel('에러발생!! 나중에 다시 시도해주세요 :)')
    driver = webdriver.Chrome(executable_path='./chromedriver', options=options)
    
    now=datetime.datetime.now()
    nowHour = now.hour
    nowMinute = now.minute
    
    nameCount = 1
    priceCount = 1
    
    
    deviceNameList = []
    lowPriceList = []
    reviewList = []
    hyperlinkList = []
    
    changePriceList = []
    changeReviewList = []
    changeHyperList = []
    
    urlShort = pyshorteners.Shortener()
    
    driver.implicitly_wait(3)
    # url에 접근한다.
    driver.get(url)

    #주간 버튼 클릭 (일간일때는 요놈만 주석처리하거나 나중에 선택지를 주는 방법으로 선택가능)
    driver.find_element_by_xpath("//*[@id='__next']/div/div[2]/div[2]/div/div[2]/div[2]/a[2]").click()


    doScrollDown(2,driver) #2초간 스크롤 넉넉하게 2초 동안 스크롤함

    req = driver.page_source #스크롤 한 후 페이지 소스 불러오기
    soup = BeautifulSoup(req, 'html.parser')

    titles = soup.select('div.imageProduct_title__3TsP1') #제품 이름 출력
    shopLink = soup.select('a.imageProduct_link_item__2i1IN') #쇼핑 링크
    lowPrices = soup.select('div.imageProduct_price__3vXjm') #최저가
    


    print('주간 Best100 크롤링 초기 데이터 :)')

    for title in titles:
    
        printTitles = title.get_text()
    
        filterTitle = filtering_string(printTitles)
        print(nameCount)
        print(printTitles)
    
        nameCount = nameCount+1
    
        deviceNameList.append(filterTitle)
    
    
    print() 
    print("순위별 최저가")   
    

    for lowPrice in lowPrices:
    
        printPrice = lowPrice.get_text()
        filterPrice = filtering_string(printPrice)
        print(priceCount)
        print(printPrice)

        priceCount = priceCount+1
    
        lowPriceList.append(int(filterPrice))
        

    reviewIndex = 1
    for links in shopLink:
        
        hyperlinkList.append(str(links['href']))
        driver.get(links["href"])
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')
        review = soup.select_one('#section_review > div.filter_sort_group__Y8HA1 > div.filter_evaluation_tap__-45pp > ul > li.filter_on__X0_Fb > a > em')
        
        if review == None:
            
            print(str(reviewIndex) +' '+'실패(-1)')
            reviewList.append(-1)
            reviewIndex = reviewIndex + 1
        else:
            reviewText = review.get_text()
            filterReviewCount = filtering_string(reviewText)
            print(str(reviewIndex) +' '+filterReviewCount)
            
            reviewList.append(int(filterReviewCount))
            reviewIndex = reviewIndex + 1
            

    
    for i in range(2, 102):
        changePriceList.append('=IF(VLOOKUP($B'+str(i)+',초기데이터!$B:$D,COLUMN(B'+str(i-1)+'),0)=C'+str(i)+',"변동없음",C'+str(i)+'-VLOOKUP($B'+str(i)+',초기데이터!$B:$D,COLUMN(B'+str(i-1)+'),0))')
        changeReviewList.append('=IF(VLOOKUP($B'+str(i)+',초기데이터!$B:$D,COLUMN(C'+str(i-1)+'),0)=D'+str(i)+',"변동없음",D'+str(i)+'-VLOOKUP($B'+str(i)+',초기데이터!$B:$D,COLUMN(C'+str(i-1)+'),0))')
        
        
    for i in range (len(hyperlinkList)):
        changeHyperList.append('=HYPERLINK("{}", "{}")'.format(urlShort.tinyurl.short(str(hyperlinkList[i])), "바로가기"))
    

    list_of_tuples = list(zip(deviceNameList,lowPriceList,reviewList,changePriceList,changeReviewList,changeHyperList))
    
    df = pd.DataFrame(list_of_tuples, columns = ['제품명', '최저가', '리뷰수','차액','리뷰 변화','바로가기'])
    
    df.index = df.index+1
    
    #df.style.format({hyperlinkList: make_clickable})
    
    
    
    with pd.ExcelWriter(nowDate+'.xlsx', mode='a', engine='openpyxl') as writer:
        df.to_excel(writer, header=True, index= True, sheet_name='최저가데이터 '+nowDate+' '+str(nowHour)+'시'+str(nowMinute)+'분',index_label='순위')
    
    workbook = load_workbook(filename=nowDate+'.xlsx')
    sheet = workbook['최저가데이터 '+nowDate+' '+str(nowHour)+'시'+str(nowMinute)+'분']
    red_fill = PatternFill(bgColor='F6C9CE')
    yellow_fill = PatternFill(bgColor='FCECA6')
    green_fill = PatternFill(bgColor='CEEED0')
    
    new_style = DifferentialStyle(fill=yellow_fill)
    review_style = DifferentialStyle(fill=green_fill)
    price_style = DifferentialStyle(fill=red_fill)
    
    rule_new = Rule(type='expression',dxf=new_style, stopIfTrue=True)
    rule_review = Rule(type='expression',dxf=review_style, stopIfTrue=True)
    rule_price = Rule(type='expression',dxf=price_style, stopIfTrue=True)
    
    rule_new.formula = ["=NOT(COUNTIF(초기데이터!$B:$B,B1)>0)"]
    rule_review.formula = ["=NOT(COUNTIF(초기데이터!$B:$D,D2)>0)"]
    rule_price.formula = ["=NOT(COUNTIF(초기데이터!$B:$C,C2)>0)"]
    
    sheet.conditional_formatting.add('C2:C101', rule_price)
    sheet.conditional_formatting.add('D2:D101', rule_review)
    sheet.conditional_formatting.add('B1:B101', rule_new)
  
    
    
    
    
    workbook.save(nowDate+'.xlsx')
    workbook.close()
    
    driver.quit()
    activeLabel('best100 완료')
    

       
def main():
    pass
    # while(1):
    #     print('네이버 쇼핑 크롤링 봇')
    #     print("1. 초기 데이터 수집")
    #     print("2. 현재 Best 순위 수집")
    #     print("0. 종료")
    #     choose = int(input('>> '))
        
        
    #     if choose == 1:            
    #         initPrototype(nameCount,priceCount)
    #     elif choose == 2:
    #         bestRank(nameCount, priceCount)
    #     elif choose == 0:
    #         driver.quit()
    #         exit()
    
    


window = tkinter.Tk()
    
window.title('네이버 쇼핑 크롤링 봇')
window.geometry('300x150+200+100')
window.resizable(False,False)
pTitle = Label(window, text = '네이버 쇼핑 크롤링 봇')

    
pProgress = Label(window, text= "버전 1.0")
    
initBtn = Button(window, text='초기데이터 수집', command=initPrototype)
bestRankBtn = Button(window, text='현재 Best 순위 수집', command=bestRank)
    
   
    
    
    
pTitle.pack()
initBtn.pack()
bestRankBtn.pack()

pProgress.pack()

    
window.mainloop()



        
    

    
    
    